# -*- coding: utf-8 -*-
"""
Servicio para exportar reportes de producción a PDF y Excel con enfoque en:
- Robustez frente a datos faltantes o tipos inesperados
- Soporte de miles de filas (tablas paginadas en PDF y write_only en Excel)
- Estilos consistentes y legibles
- Sin dependencias de campos no garantizados por el contrato
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, Any, Iterable, List, Tuple, Optional
from functools import partial
import logging

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import WriteOnlyCell

logger = logging.getLogger(__name__)


# =========================
# Utilidades generales
# =========================

def _money(x: Any) -> float:
    try:
        return float(x or 0)
    except Exception:
        return 0.0


def _pct(x: Any) -> float:
    try:
        return float(x or 0)
    except Exception:
        return 0.0


def _safe_str(x: Any) -> str:
    return "" if x is None else str(x)


def _first(s: str, n: int = 10) -> str:
    return _safe_str(s)[:n] if s else ""


def _chunk_rows(rows: List[List[Any]], size: int) -> Iterable[List[List[Any]]]:
    """Parte filas en bloques para evitar tablas gigantes en PDF."""
    for i in range(0, len(rows), size):
        yield rows[i : i + size]


def _table_style_header(bg: colors.Color) -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ]
    )


def _table_style_body(zebra: Tuple[colors.Color, colors.Color] = (colors.white, colors.lightgrey)) -> TableStyle:
    return TableStyle(
        [
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), zebra),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )


def _pdf_doc(buffer: BytesIO) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=0.75 * inch,   # refinado
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        bottomMargin=0.75 * inch,
    )


def _pdf_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=20,
        alignment=1,
        textColor=HexColor("#2e7d32"),
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=10,
        textColor=HexColor("#0288d1"),
    )
    note_style = ParagraphStyle(
        "Note",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.grey,
        spaceBefore=4,
        spaceAfter=4,
    )
    return title_style, heading_style, note_style


def _pdf_header_footer(canvas, doc, title: str):
    canvas.saveState()
    width, height = A4
    canvas.setFillColor(HexColor("#2e7d32"))
    canvas.rect(0, height - 50, width, 50, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawString(40, height - 30, title)
    canvas.setFont("Helvetica", 9)
    canvas.setFillColor(colors.grey)
    canvas.drawString(40, 30, "Agroproductores Risol")
    canvas.drawRightString(width - 40, 30, f"Página {doc.page}")
    canvas.restoreState()


# =========================
# Utilidades para Excel (write_only-safe)
# =========================

def _woc(ws: Worksheet, value, font: Optional[Font] = None, fill: Optional[PatternFill] = None, number_format: Optional[str] = None):
    """Crea una WriteOnlyCell con formato aplicado antes de append (requerido en write_only=True)."""
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    return c


def _ws_header(ws: Worksheet, title: str, merge_to_col: int = 6):
    """
    Cabecera de hoja compatible con write_only:
    - En write_only: agrega una fila con la celda A1 estilizada (sin merge).
    - En modo normal: escribe en A1 y realiza merge opcional.
    """
    is_write_only = getattr(ws.parent, "write_only", False)
    if is_write_only:
        title_cell = _woc(ws, title, font=Font(bold=True, size=16))
        # Rellenar columnas para mantener ancho lógico
        pad = [WriteOnlyCell(ws, value="") for _ in range(max(0, merge_to_col - 1))]
        ws.append([title_cell] + pad)
    else:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16)
        if merge_to_col and merge_to_col > 1:
            end_col_letter = chr(ord("A") + merge_to_col - 1)
            ws.merge_cells(f"A1:{end_col_letter}1")


def _ws_section_title(ws: Worksheet, row: int, text: str):
    """
    Título de sección compatible con write_only.
    - En write_only: agrega fila con fondo y texto blanco (sin coordenadas).
    - En modo normal: escribe en A{row}.
    """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    is_write_only = getattr(ws.parent, "write_only", False)
    if is_write_only:
        ws.append([_woc(ws, text, font=header_font, fill=header_fill)])
    else:
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill


def _apply_border(ws: Worksheet, cell_range: str):
    # Nota: no usar en write_only. Mantener para hojas normales.
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


# ===================================================
# Integración opcional con WeasyPrint (reporting.py)
# ===================================================

def _try_weasy_from_reporting(kind: str, reporte_data: Dict[str, Any]) -> Optional[bytes]:
    """
    Intenta renderizar el PDF con los templates de reporting (WeasyPrint).
    Si falla por cualquier razón, devuelve None y se usa ReportLab como fallback.
    """
    try:
        # Import perezoso para no forzar dependencia si no se usa
        from gestion_huerta.utils import reporting as rep  # type: ignore
    except Exception:
        return None

    meta = (reporte_data or {}).get("metadata", {}) or {}
    try:
        if kind == "cosecha" and meta.get("cosecha_id"):
            return rep.render_cosecha_pdf(int(meta["cosecha_id"]))
        if kind == "temporada" and meta.get("temporada_id"):
            return rep.render_temporada_pdf(int(meta["temporada_id"]))
        if kind == "perfil_huerta":
            # reporting.render_huerta_pdf espera un ID (propia o rentada)
            hid = meta.get("huerta_id") or meta.get("huerta_rentada_id")
            if hid:
                return rep.render_huerta_pdf(int(hid))
        return None
    except Exception as e:
        logger.debug("WeasyPrint fallback (%s) -> ReportLab. Motivo: %s", kind, e)
        return None


# ===================================================
# Servicio principal de exportación (PDF / Excel)
# ===================================================

class ExportacionService:
    """Servicio para exportar reportes a PDF y Excel (cosecha, temporada, perfil huerta)."""

    # =========================
    # COSECHA
    # =========================

    @staticmethod
    def generar_pdf_cosecha(reporte_data: Dict[str, Any]) -> bytes:
        """
        Genera PDF del reporte de cosecha.
        Intenta WeasyPrint (reporting) y hace fallback a ReportLab si no está disponible.
        """
        weasy = _try_weasy_from_reporting("cosecha", reporte_data)
        if weasy:
            return weasy

        buffer = BytesIO()
        doc = _pdf_doc(buffer)
        title_style, heading_style, note_style = _pdf_styles()

        story: List[Any] = []
        story.append(Paragraph("REPORTE DE COSECHA", title_style))  # refinado

        info = reporte_data.get("informacion_general", {}) or {}
        resumen = reporte_data.get("resumen_financiero", {}) or {}

        # Información General
        story.append(Paragraph("INFORMACIÓN GENERAL", heading_style))
        info_data = [
            ["Huerta:", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"],
            ["Propietario:", _safe_str(info.get("propietario"))],
            ["Temporada:", _safe_str(info.get("temporada_año"))],
            ["Cosecha:", _safe_str(info.get("cosecha_nombre"))],
            ["Estado:", _safe_str(info.get("estado"))],
            ["Hectáreas:", f"{_money(info.get('hectareas')):.2f} ha"],
        ]
        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch], repeatRows=0, hAlign="LEFT")
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(info_table)
        story.append(Spacer(1, 10))

        # Resumen financiero
        story.append(Paragraph("RESUMEN FINANCIERO", heading_style))
        resumen_data = [
            ["Concepto", "Monto"],
            ["Total Inversiones", f"${_money(resumen.get('total_inversiones')):,.2f}"],
            ["Total Ventas", f"${_money(resumen.get('total_ventas')):,.2f}"],
            ["Gastos de Venta", f"${_money(resumen.get('total_gastos_venta')):,.2f}"],
            ["Ganancia Bruta", f"${_money(resumen.get('ganancia_bruta')):,.2f}"],
            ["Ganancia Neta", f"${_money(resumen.get('ganancia_neta')):,.2f}"],
            ["ROI", f"{_pct(resumen.get('roi_porcentaje')):.1f}%"],
            ["Ganancia por Hectárea", f"${_money(resumen.get('ganancia_por_hectarea')):,.2f}"],
        ]
        res_table = Table(resumen_data, colWidths=[3 * inch, 2.2 * inch], repeatRows=1, hAlign="LEFT")
        res_table.setStyle(_table_style_header(HexColor("#2e7d32")))
        res_table.setStyle(_table_style_body())
        story.append(res_table)
        story.append(Spacer(1, 10))

        # Detalle de Inversiones (paginado)
        inv_headers = ["Fecha", "Categoría", "Insumos", "Mano Obra", "Total", "Descripción"]
        inv_rows: List[List[Any]] = [inv_headers]
        for inv in reporte_data.get("detalle_inversiones", []) or []:
            inv_rows.append(
                [
                    _first(inv.get("fecha")),
                    _safe_str(inv.get("categoria") or "Sin categoría"),
                    f"${_money(inv.get('gastos_insumos')):,.2f}",
                    f"${_money(inv.get('gastos_mano_obra')):,.2f}",
                    f"${_money(inv.get('total')):,.2f}",
                    _safe_str(inv.get("descripcion")),
                ]
            )

        story.append(Paragraph("DETALLE DE INVERSIONES", heading_style))
        if len(inv_rows) == 1:
            story.append(Paragraph("Sin inversiones registradas.", note_style))
        else:
            header = inv_rows[0]
            body = inv_rows[1:]
            for i, chunk in enumerate(_chunk_rows(body, 800)):  # 800 filas por tabla ≈ seguro en A4
                data = [header] + chunk
                t = Table(
                    data,
                    colWidths=[1.0 * inch, 1.5 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 1.4 * inch],
                    repeatRows=1,
                    splitByRow=True,
                )
                t.setStyle(_table_style_header(HexColor("#0288d1")))
                t.setStyle(_table_style_body())
                # refinado: alinear texto libre a la izquierda
                t.setStyle(TableStyle([
                    ("ALIGN", (1, 1), (1, -1), "LEFT"),   # Categoría
                    ("ALIGN", (5, 1), (5, -1), "LEFT"),   # Descripción
                ]))
                story.append(t)
                if i < (len(body) - 1) // 800:
                    story.append(PageBreak())

        story.append(Spacer(1, 10))

        # Detalle de Ventas (paginado)
        ven_headers = ["Fecha", "Variedad", "Cajas", "Precio/Caja", "Total", "Gasto", "Utilidad Neta"]
        ven_rows: List[List[Any]] = [ven_headers]
        for v in reporte_data.get("detalle_ventas", []) or []:
            ingreso = _money(v.get("total_venta"))
            gasto = _money(v.get("gasto"))
            util = _money(v.get("ganancia_neta")) if "ganancia_neta" in v else (ingreso - gasto)
            ven_rows.append(
                [
                    _first(v.get("fecha")),
                    _safe_str(v.get("tipo_mango")),
                    str(int(v.get("num_cajas") or 0)),
                    f"${_money(v.get('precio_por_caja')):,.2f}",
                    f"${ingreso:,.2f}",
                    f"${gasto:,.2f}",
                    f"${util:,.2f}",
                ]
            )

        story.append(Paragraph("DETALLE DE VENTAS", heading_style))
        if len(ven_rows) == 1:
            story.append(Paragraph("Sin ventas registradas.", note_style))
        else:
            header = ven_rows[0]
            body = ven_rows[1:]
            for i, chunk in enumerate(_chunk_rows(body, 800)):
                data = [header] + chunk
                t = Table(
                    data,
                    colWidths=[1.0 * inch, 1.5 * inch, 0.8 * inch, 1.1 * inch, 1.2 * inch, 1.0 * inch, 1.2 * inch],
                    repeatRows=1,
                    splitByRow=True,
                )
                t.setStyle(_table_style_header(HexColor("#FF8C00")))
                t.setStyle(_table_style_body())
                # refinado: Variedad a la izquierda
                t.setStyle(TableStyle([
                    ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ]))
                story.append(t)
                if i < (len(body) - 1) // 800:
                    story.append(PageBreak())

        # (Opcional) Análisis
        cats = reporte_data.get("analisis_categorias") or []
        if cats:
            story.append(Spacer(1, 10))
            story.append(Paragraph("ANÁLISIS DE CATEGORÍAS (Inversiones)", heading_style))
            data = [["Categoría", "Total", "%"]]
            for c in cats:
                data.append(
                    [
                        _safe_str(c.get("categoria")),
                        f"${_money(c.get('total')):,.2f}",
                        f"{_pct(c.get('porcentaje')):.1f}%",
                    ]
                )
            t = Table(data, colWidths=[2.5 * inch, 1.6 * inch, 0.8 * inch], repeatRows=1)
            t.setStyle(_table_style_header(HexColor("#2F5597")))
            t.setStyle(_table_style_body())
            story.append(t)

        vars_ = reporte_data.get("analisis_variedades") or []
        if vars_:
            story.append(Spacer(1, 10))
            story.append(Paragraph("ANÁLISIS DE VARIEDADES (Ventas)", heading_style))
            data = [["Variedad", "Cajas", "Precio Prom.", "Total", "%"]]
            for r in vars_:
                data.append(
                    [
                        _safe_str(r.get("variedad")),
                        str(int(r.get("total_cajas") or 0)),
                        f"${_money(r.get('precio_promedio')):,.2f}",
                        f"${_money(r.get('total_venta')):,.2f}",
                        f"{_pct(r.get('porcentaje')):.1f}%",
                    ]
                )
            t = Table(data, colWidths=[2.0 * inch, 0.9 * inch, 1.2 * inch, 1.6 * inch, 0.8 * inch], repeatRows=1)
            t.setStyle(_table_style_header(HexColor("#7030A0")))
            t.setStyle(_table_style_body())
            story.append(t)

        header = partial(_pdf_header_footer, title="Reporte de Cosecha")
        doc.build(story, onFirstPage=header, onLaterPages=header)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generar_excel_cosecha(reporte_data: Dict[str, Any]) -> bytes:
        """
        Genera Excel del reporte de cosecha.
        Usa `Workbook(write_only=True)` para escalar a miles de filas sin agotar memoria.
        *Compatibilidad write_only*: no se usan índices de celda ni iter_rows; todo via WriteOnlyCell.
        """
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

        # Hoja Resumen
        ws_resumen = wb.create_sheet("Resumen")
        _ws_header(ws_resumen, "REPORTE DE PRODUCCIÓN - COSECHA INDIVIDUAL", merge_to_col=6)

        info = reporte_data.get("informacion_general", {}) or {}
        resumen = reporte_data.get("resumen_financiero", {}) or {}

        row = 3
        _ws_section_title(ws_resumen, row, "INFORMACIÓN GENERAL")
        info_data = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Temporada", _safe_str(info.get("temporada_año"))),
            ("Cosecha", _safe_str(info.get("cosecha_nombre"))),
            ("Estado", _safe_str(info.get("estado"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
        ]
        for label, value in info_data:
            row += 1
            ws_resumen.append([_woc(ws_resumen, label, font=Font(bold=True)), _woc(ws_resumen, value)])

        row += 2
        _ws_section_title(ws_resumen, row, "RESUMEN FINANCIERO")
        resumen_rows = [
            ("Total Inversiones", _money(resumen.get("total_inversiones")), "#,##0.00"),
            ("Total Ventas", _money(resumen.get("total_ventas")), "#,##0.00"),
            ("Gastos de Venta", _money(resumen.get("total_gastos_venta")), "#,##0.00"),
            ("Ganancia Bruta", _money(resumen.get("ganancia_bruta")), "#,##0.00"),
            ("Ganancia Neta", _money(resumen.get("ganancia_neta")), "#,##0.00"),
            ("ROI (%)", _pct(resumen.get("roi_porcentaje")), "0.0"),
            ("Ganancia por Hectárea", _money(resumen.get("ganancia_por_hectarea")), "#,##0.00"),
        ]
        for label, value, fmt in resumen_rows:
            ws_resumen.append([_woc(ws_resumen, label, font=Font(bold=True)), _woc(ws_resumen, value, number_format=fmt)])

        ws_resumen.freeze_panes = "A4"

        # Hoja Inversiones
        ws_inv = wb.create_sheet("Inversiones")
        ws_inv.append([
            _woc(ws_inv, "Fecha",      font=header_font, fill=blue_fill),
            _woc(ws_inv, "Categoría",  font=header_font, fill=blue_fill),
            _woc(ws_inv, "Insumos",    font=header_font, fill=blue_fill),
            _woc(ws_inv, "Mano Obra",  font=header_font, fill=blue_fill),
            _woc(ws_inv, "Total",      font=header_font, fill=blue_fill),
            _woc(ws_inv, "Descripción",font=header_font, fill=blue_fill),
        ])
        for inv in reporte_data.get("detalle_inversiones", []) or []:
            ws_inv.append([
                _woc(ws_inv, _first(inv.get("fecha"))),
                _woc(ws_inv, _safe_str(inv.get("categoria") or "Sin categoría")),
                _woc(ws_inv, _money(inv.get("gastos_insumos")),   number_format="#,##0.00"),
                _woc(ws_inv, _money(inv.get("gastos_mano_obra")), number_format="#,##0.00"),
                _woc(ws_inv, _money(inv.get("total")),            number_format="#,##0.00"),
                _woc(ws_inv, _safe_str(inv.get("descripcion"))),
            ])
        ws_inv.freeze_panes = "A2"

        # Hoja Ventas
        ws_ven = wb.create_sheet("Ventas")
        ws_ven.append([
            _woc(ws_ven, "Fecha",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Variedad",     font=header_font, fill=purple_fill),
            _woc(ws_ven, "Cajas",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Precio/Caja",  font=header_font, fill=purple_fill),
            _woc(ws_ven, "Total",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Gasto",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Utilidad Neta",font=header_font, fill=purple_fill),
        ])
        for v in reporte_data.get("detalle_ventas", []) or []:
            ingreso = _money(v.get("total_venta"))
            gasto   = _money(v.get("gasto"))
            util    = _money(v.get("ganancia_neta")) if "ganancia_neta" in v else (ingreso - gasto)
            ws_ven.append([
                _woc(ws_ven, _first(v.get("fecha"))),
                _woc(ws_ven, _safe_str(v.get("tipo_mango"))),
                _woc(ws_ven, int(v.get("num_cajas") or 0),       number_format="0"),
                _woc(ws_ven, _money(v.get("precio_por_caja")),   number_format="#,##0.00"),
                _woc(ws_ven, ingreso,                             number_format="#,##0.00"),
                _woc(ws_ven, gasto,                               number_format="#,##0.00"),
                _woc(ws_ven, util,                                number_format="#,##0.00"),
            ])
        ws_ven.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================
    # TEMPORADA
    # =========================

    @staticmethod
    def generar_pdf_temporada(reporte_data: Dict[str, Any]) -> bytes:
        """
        Genera PDF de temporada: intenta WeasyPrint (reporting) y cae a ReportLab si no.
        """
        weasy = _try_weasy_from_reporting("temporada", reporte_data)
        if weasy:
            return weasy

        buffer = BytesIO()
        doc = _pdf_doc(buffer)
        title_style, heading_style, note_style = _pdf_styles()

        story: List[Any] = []
        story.append(Paragraph("REPORTE DE TEMPORADA", title_style))  # refinado

        info = reporte_data.get("informacion_general", {}) or {}
        res = reporte_data.get("resumen_ejecutivo", {}) or {}

        story.append(Paragraph("INFORMACIÓN GENERAL", heading_style))
        info_data = [
            ["Huerta:", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"],
            ["Propietario:", _safe_str(info.get("propietario"))],
            ["Temporada:", _safe_str(info.get("temporada_año"))],
            ["Total Cosechas:", _safe_str(info.get("total_cosechas"))],
            ["Estado:", _safe_str(info.get("estado"))],
            ["Hectáreas:", f"{_money(info.get('hectareas')):.2f} ha"],
        ]
        t_info = Table(info_data, colWidths=[2 * inch, 4 * inch])
        t_info.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(t_info)
        story.append(Spacer(1, 10))

        story.append(Paragraph("RESUMEN EJECUTIVO", heading_style))
        res_data = [
            ["Concepto", "Valor"],
            ["Inversión Total", f"${_money(res.get('inversion_total')):,.2f}"],
            ["Ventas Totales", f"${_money(res.get('ventas_totales')):,.2f}"],
            ["Ganancia Neta", f"${_money(res.get('ganancia_neta')):,.2f}"],
            ["ROI Temporada", f"{_pct(res.get('roi_temporada')):.1f}%"],
            ["Productividad", f"{_money(res.get('productividad')):.1f} cajas/ha"],
            ["Cajas Totales", _safe_str(res.get("cajas_totales"))],
        ]
        t_res = Table(res_data, colWidths=[3 * inch, 2.2 * inch], repeatRows=1)
        t_res.setStyle(_table_style_header(HexColor("#2e7d32")))
        t_res.setStyle(_table_style_body())
        story.append(t_res)
        story.append(Spacer(1, 10))

        story.append(Paragraph("COMPARATIVO POR COSECHA", heading_style))
        comp_headers = ["Cosecha", "Inversión", "Ventas", "Ganancia", "ROI", "Cajas"]
        comp_rows = [comp_headers]
        for c in reporte_data.get("comparativo_cosechas", []) or []:
            comp_rows.append(
                [
                    _safe_str(c.get("nombre")),
                    f"${_money(c.get('inversion')):,.2f}",
                    f"${_money(c.get('ventas')):,.2f}",
                    f"${_money(c.get('ganancia')):,.2f}",
                    f"{_pct(c.get('roi')):.1f}%",
                    _safe_str(c.get("cajas")),
                ]
            )

        if len(comp_rows) == 1:
            story.append(Paragraph("Sin cosechas activas.", note_style))
        else:
            header = comp_rows[0]
            body = comp_rows[1:]
            for i, chunk in enumerate(_chunk_rows(body, 800)):
                t = Table(
                    [header] + chunk,
                    colWidths=[1.6 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 0.9 * inch, 0.8 * inch],
                    repeatRows=1,
                    splitByRow=True,
                )
                t.setStyle(_table_style_header(HexColor("#0288d1")))
                t.setStyle(_table_style_body())
                # refinado: nombre de cosecha a la izquierda
                t.setStyle(TableStyle([
                    ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ]))
                story.append(t)
                if i < (len(body) - 1) // 800:
                    story.append(PageBreak())

        header_fn = partial(_pdf_header_footer, title="Reporte de Temporada")
        doc.build(story, onFirstPage=header_fn, onLaterPages=header_fn)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generar_excel_temporada(reporte_data: Dict[str, Any]) -> bytes:
        """
        Excel de temporada en modo write_only (formato aplicado al crear cada celda).
        """
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        ws = wb.create_sheet("Resumen Ejecutivo")
        _ws_header(ws, "REPORTE DE PRODUCCIÓN - TEMPORADA COMPLETA", merge_to_col=6)

        info = reporte_data.get("informacion_general", {}) or {}
        res = reporte_data.get("resumen_ejecutivo", {}) or {}

        row = 3
        _ws_section_title(ws, row, "INFORMACIÓN GENERAL")
        info_rows = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Temporada", _safe_str(info.get("temporada_año"))),
            ("Total Cosechas", _safe_str(info.get("total_cosechas"))),
            ("Estado", _safe_str(info.get("estado"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
        ]
        for label, value in info_rows:
            row += 1
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value)])

        row += 2
        _ws_section_title(ws, row, "RESUMEN EJECUTIVO")
        res_rows = [
            ("Inversión Total", _money(res.get("inversion_total")), "#,##0.00"),
            ("Ventas Totales", _money(res.get("ventas_totales")), "#,##0.00"),
            ("Ganancia Neta", _money(res.get("ganancia_neta")), "#,##0.00"),
            ("ROI Temporada (%)", _pct(res.get("roi_temporada")), "0.0"),
            ("Productividad (cajas/ha)", _money(res.get("productividad")), "#,##0.00"),
            ("Cajas Totales", _money(res.get("cajas_totales")), "0"),  # refinado
        ]
        for label, value, fmt in res_rows:
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value, number_format=fmt)])
        ws.freeze_panes = "A4"

        ws_comp = wb.create_sheet("Comparativo Cosechas")
        ws_comp.append([
            _woc(ws_comp, "Cosecha",  font=header_font, fill=blue_fill),
            _woc(ws_comp, "Inversión",font=header_font, fill=blue_fill),
            _woc(ws_comp, "Ventas",   font=header_font, fill=blue_fill),
            _woc(ws_comp, "Ganancia", font=header_font, fill=blue_fill),
            _woc(ws_comp, "ROI (%)",  font=header_font, fill=blue_fill),
            _woc(ws_comp, "Cajas",    font=header_font, fill=blue_fill),
        ])
        for c in (reporte_data.get("comparativo_cosechas") or []):
            ws_comp.append([
                _woc(ws_comp, _safe_str(c.get("nombre"))),
                _woc(ws_comp, _money(c.get("inversion")), number_format="#,##0.00"),
                _woc(ws_comp, _money(c.get("ventas")),    number_format="#,##0.00"),
                _woc(ws_comp, _money(c.get("ganancia")),  number_format="#,##0.00"),
                _woc(ws_comp, _pct(c.get("roi")),         number_format="0.0"),
                _woc(ws_comp, _money(c.get("cajas")),     number_format="0"),  # refinado
            ])
        ws_comp.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================
    # PERFIL HUERTA
    # =========================

    @staticmethod
    def generar_pdf_perfil_huerta(reporte_data: Dict[str, Any]) -> bytes:
        """
        Genera PDF de perfil de huerta: intenta WeasyPrint y cae a ReportLab si no.
        """
        weasy = _try_weasy_from_reporting("perfil_huerta", reporte_data)
        if weasy:
            return weasy

        buffer = BytesIO()
        doc = _pdf_doc(buffer)
        title_style, heading_style, note_style = _pdf_styles()

        story: List[Any] = []
        story.append(Paragraph("PERFIL DE HUERTA", title_style))  # refinado

        info = reporte_data.get("informacion_general", {}) or {}

        story.append(Paragraph("INFORMACIÓN GENERAL", heading_style))
        info_data = [
            ["Huerta:", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"],
            ["Propietario:", _safe_str(info.get("propietario"))],
            ["Ubicación:", _safe_str(info.get("ubicacion"))],
            ["Hectáreas:", f"{_money(info.get('hectareas')):.2f} ha"],
            ["Variedades:", _safe_str(info.get("variedades"))],
            ["Años de Operación:", _safe_str(info.get("años_operacion"))],
            ["Temporadas Analizadas:", _safe_str(info.get("temporadas_analizadas"))],
        ]
        t_info = Table(info_data, colWidths=[2 * inch, 4 * inch])
        t_info.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(t_info)
        story.append(Spacer(1, 10))

        # Resumen histórico (paginado)
        story.append(Paragraph("RESUMEN HISTÓRICO", heading_style))
        hist_headers = ["Año", "Inversión", "Ventas", "Ganancia", "ROI", "Productividad", "Cosechas"]
        hist_rows = [hist_headers]
        for d in reporte_data.get("resumen_historico", []) or []:
            hist_rows.append(
                [
                    _safe_str(d.get("año")),
                    f"${_money(d.get('inversion')):,.2f}",
                    f"${_money(d.get('ventas')):,.2f}",
                    f"${_money(d.get('ganancia')):,.2f}",
                    f"{_pct(d.get('roi')):.1f}%",
                    f"{_money(d.get('productividad'))} cajas",
                    _safe_str(d.get("cosechas_count")),
                ]
            )
        if len(hist_rows) == 1:
            story.append(Paragraph("Sin datos históricos.", note_style))
        else:
            header = hist_rows[0]
            body = hist_rows[1:]
            for i, chunk in enumerate(_chunk_rows(body, 800)):
                t = Table(
                    [header] + chunk,
                    colWidths=[0.8 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 0.8 * inch, 1.2 * inch, 0.8 * inch],
                    repeatRows=1,
                    splitByRow=True,
                )
                t.setStyle(_table_style_header(HexColor("#2e7d32")))
                t.setStyle(_table_style_body())
                story.append(t)
                if i < (len(body) - 1) // 800:
                    story.append(PageBreak())

        # Análisis de eficiencia (si existe)
        ef = reporte_data.get("analisis_eficiencia") or {}
        story.append(Spacer(1, 10))
        story.append(Paragraph("ANÁLISIS DE EFICIENCIA", heading_style))
        ef_data = [
            ["Métrica", "Valor"],
            [
                "Mejor Temporada",
                f"{_safe_str(ef.get('mejor_temporada', {}).get('año'))} "
                f"(ROI: {_pct(ef.get('mejor_temporada', {}).get('roi')):.1f}%)",
            ],
            [
                "Peor Temporada",
                f"{_safe_str(ef.get('peor_temporada', {}).get('año'))} "
                f"(ROI: {_pct(ef.get('peor_temporada', {}).get('roi')):.1f}%)",
            ],
            ["ROI Promedio Histórico", f"{_pct(ef.get('roi_promedio_historico')):.1f}%"],
            ["Variabilidad ROI", f"±{_pct(ef.get('variabilidad_roi')):.1f}%"],
            ["Tendencia", _safe_str(ef.get("tendencia"))],
        ]
        t_ef = Table(ef_data, colWidths=[3 * inch, 3 * inch], repeatRows=1)
        t_ef.setStyle(_table_style_header(HexColor("#0288d1")))
        t_ef.setStyle(_table_style_body())
        story.append(t_ef)

        # Proyecciones (si existen)
        pr = reporte_data.get("proyecciones") or {}
        story.append(Spacer(1, 10))
        story.append(Paragraph("PROYECCIONES", heading_style))
        pr_data = [
            ["Proyección de Ventas Próx. Temporada", f"${_money(pr.get('proyeccion_proxima_temporada')):,.2f}"],
            ["ROI Esperado", f"{_pct(pr.get('roi_esperado')):.1f}%"],
            ["Recomendaciones", ", ".join([_safe_str(x) for x in (pr.get("recomendaciones") or [])]) or "-"],
            ["Alertas", ", ".join([_safe_str(x) for x in (pr.get("alertas") or [])]) or "-"],
        ]
        t_pr = Table(pr_data, colWidths=[3 * inch, 3 * inch], repeatRows=1)
        t_pr.setStyle(_table_style_header(HexColor("#9E480E")))
        t_pr.setStyle(_table_style_body())
        story.append(t_pr)

        header_fn = partial(_pdf_header_footer, title="Perfil de Huerta")
        doc.build(story, onFirstPage=header_fn, onLaterPages=header_fn)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generar_excel_perfil_huerta(reporte_data: Dict[str, Any]) -> bytes:
        """
        Excel de perfil de huerta en modo write_only, sin merges ni indexaciones directas.
        """
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        orange_fill = PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid")

        ws = wb.create_sheet("Resumen Histórico")
        _ws_header(ws, "PERFIL HISTÓRICO DE HUERTA", merge_to_col=7)

        info = reporte_data.get("informacion_general", {}) or {}

        row = 3
        _ws_section_title(ws, row, "INFORMACIÓN GENERAL")
        info_rows = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Ubicación", _safe_str(info.get("ubicacion"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
            ("Variedades", _safe_str(info.get("variedades"))),
            ("Años de Operación", _safe_str(info.get("años_operacion"))),
            ("Temporadas Analizadas", _safe_str(info.get("temporadas_analizadas"))),
        ]
        for label, value in info_rows:
            row += 1
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value)])

        row += 2
        _ws_section_title(ws, row, "RESUMEN HISTÓRICO")
        ws.append([
            _woc(ws, "Año",           font=header_font, fill=blue_fill),
            _woc(ws, "Inversión",     font=header_font, fill=blue_fill),
            _woc(ws, "Ventas",        font=header_font, fill=blue_fill),
            _woc(ws, "Ganancia",      font=header_font, fill=blue_fill),
            _woc(ws, "ROI (%)",       font=header_font, fill=blue_fill),
            _woc(ws, "Productividad", font=header_font, fill=blue_fill),
            _woc(ws, "Cosechas",      font=header_font, fill=blue_fill),
        ])
        for d in (reporte_data.get("resumen_historico") or []):
            ws.append([
                _woc(ws, _safe_str(d.get("año"))),
                _woc(ws, _money(d.get("inversion")),     number_format="#,##0.00"),
                _woc(ws, _money(d.get("ventas")),        number_format="#,##0.00"),
                _woc(ws, _money(d.get("ganancia")),      number_format="#,##0.00"),
                _woc(ws, _pct(d.get("roi")),             number_format="0.0"),
                _woc(ws, _money(d.get("productividad")), number_format="#,##0.00"),
                _woc(ws, _safe_str(d.get("cosechas_count"))),
            ])
        ws.freeze_panes = "A5"

        # Hoja Proyecciones (si existen)
        pr = reporte_data.get("proyecciones") or {}
        ws2 = wb.create_sheet("Proyecciones")
        _ws_header(ws2, "Proyecciones")
        ws2.append([
            _woc(ws2, "Métrica", font=header_font, fill=orange_fill),
            _woc(ws2, "Valor",   font=header_font, fill=orange_fill),
        ])
        ws2.append([
            _woc(ws2, "Proyección Ventas Próx. Temporada"),
            _woc(ws2, _money(pr.get("proyeccion_proxima_temporada")), number_format="#,##0.00")
        ])
        ws2.append([
            _woc(ws2, "ROI Esperado (%)"),
            _woc(ws2, _pct(pr.get("roi_esperado")), number_format="0.0")
        ])
        ws2.append([
            _woc(ws2, "Recomendaciones"),
            _woc(ws2, ", ".join([_safe_str(x) for x in (pr.get("recomendaciones") or [])]) or "-")
        ])
        ws2.append([
            _woc(ws2, "Alertas"),
            _woc(ws2, ", ".join([_safe_str(x) for x in (pr.get("alertas") or [])]) or "-")
        ])
        ws2.freeze_panes = "A3"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
