# -*- coding: utf-8 -*-
from __future__ import annotations

from io import BytesIO
from typing import Dict, Any, List, Optional
from datetime import datetime
try:
    from django.utils import timezone
except Exception:
    timezone = None  # type: ignore

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import WriteOnlyCell

# =========================
# Utilidades Excel (write_only-safe)
# =========================
def _money(x: Any) -> float:
    try:
        return float(x or 0)
    except Exception:
        return 0.0

def _pct(x: Any) -> float:
    try:
        return float(x or 0)
    except Exception:
        return 0.0

def _safe_str(x: Any) -> str:
    return "" if x is None else str(x)

def _first(s: str, n: int = 10) -> str:
    return _safe_str(s)[:n] if s else ""

def _woc(
    ws: Worksheet,
    value,
    font: Optional[Font] = None,
    fill: Optional[PatternFill] = None,
    number_format: Optional[str] = None,
    alignment: Optional[Alignment] = None,
):
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    if alignment:
        c.alignment = alignment
    return c

def _ws_header(ws: Worksheet, title: str, merge_to_col: int = 6):
    """
    Escribe el título en A1 con soporte para write_only.
    En write_only no se puede mergear, así que emulamos con celdas vacías.
    """
    is_write_only = getattr(ws.parent, "write_only", False)
    if is_write_only:
        title_cell = _woc(ws, title, font=Font(bold=True, size=16))
        pad = [WriteOnlyCell(ws, value="") for _ in range(max(0, merge_to_col - 1))]
        ws.append([title_cell] + pad)
    else:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16)
        if merge_to_col and merge_to_col > 1:
            end_col_letter = chr(ord("A") + merge_to_col - 1)
            ws.merge_cells(f"A1:{end_col_letter}1")

def _ws_section_title(ws: Worksheet, row: int, text: str):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    is_write_only = getattr(ws.parent, "write_only", False)
    if is_write_only:
        ws.append([_woc(ws, text, font=header_font, fill=header_fill)])
    else:
        ws[f"A{row}"] = text
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill

def _apply_border(ws: Worksheet, cell_range: str):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def _set_column_widths(ws: Worksheet, widths: List[float]):
    """
    Ajusta anchos de columnas A.. según la lista widths (en caracteres).
    Es compatible con write_only.
    """
    for i, w in enumerate(widths, start=1):
        col_letter = chr(ord("A") + i - 1)
        ws.column_dimensions[col_letter].width = w

def _add_metadata_sheet(wb: Workbook, reporte_data: Dict[str, Any], titulo: str):
    ws_meta = wb.create_sheet("Información del reporte")
    ws_meta.append(["Título", titulo])
    gen = (
        timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")
        if timezone is not None else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    ws_meta.append(["Generado", gen])
    meta = (reporte_data or {}).get("metadata", {}) or {}
    for k, v in meta.items():
        if k in {"temporada_id", "version", "cosecha_id", "huerta_id", "huerta_rentada_id", "entidad", "infoHuerta", "fecha_generacion"}:
            continue
        ws_meta.append([_safe_str(k), _safe_str(v)])
    _set_column_widths(ws_meta, [28, 72])

# =========================
# Exportador Excel por entidad
# =========================
class ExcelExporter:
    @staticmethod
    def generar_excel_cosecha(reporte_data: Dict[str, Any]) -> bytes:
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

        # Resumen
        ws_resumen = wb.create_sheet("Resumen")
        _ws_header(ws_resumen, "REPORTE DE PRODUCCIÓN - COSECHA INDIVIDUAL", merge_to_col=6)

        info = reporte_data.get("informacion_general", {}) or {}
        resumen = reporte_data.get("resumen_financiero", {}) or {}

        # En COSECHA sí mostramos período
        fi = _first(info.get("fecha_inicio")); ff = _first(info.get("fecha_fin"))
        periodo_txt = _safe_str(info.get("periodo") or (f"{fi} - {ff}" if (fi or ff) else ""))

        row = 3
        _ws_section_title(ws_resumen, row, "INFORMACIÓN GENERAL")
        info_data = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Ubicación", _safe_str(info.get("ubicacion"))),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Temporada", _safe_str(info.get("temporada_año"))),
            ("Cosecha", _safe_str(info.get("cosecha_nombre"))),
            ("Período", periodo_txt),
            ("Estado", _safe_str(info.get("estado"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
        ]
        for label, value in info_data:
            row += 1
            ws_resumen.append([_woc(ws_resumen, label, font=Font(bold=True)), _woc(ws_resumen, value)])

        row += 2
        _ws_section_title(ws_resumen, row, "RESUMEN FINANCIERO")
        resumen_rows = [
            ("Total Inversiones", _money(resumen.get("total_inversiones")), "#,##0.00"),
            ("Total Ventas", _money(resumen.get("total_ventas")), "#,##0.00"),
            ("Gastos de Venta", _money(resumen.get("total_gastos_venta")), "#,##0.00"),
            ("Ganancia Bruta", _money(resumen.get("ganancia_bruta")), "#,##0.00"),
            ("Ganancia Neta", _money(resumen.get("ganancia_neta")), "#,##0.00"),
            ("ROI (%)", _pct(resumen.get("roi_porcentaje")), "0.0"),
            ("Ganancia por Hectárea", _money(resumen.get("ganancia_por_hectarea")), "#,##0.00"),
        ]
        for label, value, fmt in resumen_rows:
            ws_resumen.append([_woc(ws_resumen, label, font=Font(bold=True)), _woc(ws_resumen, value, number_format=fmt)])

        ws_resumen.freeze_panes = "A4"
        # Autoajuste básico para hoja Resumen (2 columnas: etiqueta / valor)
        try:
            labels = [
                "Huerta", "Ubicación", "Propietario", "Temporada", "Cosecha", "Período", "Estado", "Hectáreas",
                "Total Inversiones", "Total Ventas", "Gastos de Venta", "Ganancia Bruta", "Ganancia Neta",
                "ROI (%)", "Ganancia por Hectárea",
            ]
            col1 = max(12, max(len(_safe_str(x)) for x in labels))
            # aproximar longitud de los valores para ancho de columna 2
            val_samples = [
                _safe_str(info.get("huerta_nombre")), _safe_str(info.get("ubicacion")), _safe_str(info.get("propietario")),
                _safe_str(info.get("temporada_año")), _safe_str(info.get("cosecha_nombre")), periodo_txt, _safe_str(info.get("estado")),
                f"{_money(info.get('hectareas')):,.2f} ha",
                f"{_money(resumen.get('total_inversiones')):,.2f}",
                f"{_money(resumen.get('total_ventas')):,.2f}",
                f"{_money(resumen.get('total_gastos_venta')):,.2f}",
                f"{_money(resumen.get('ganancia_bruta')):,.2f}",
                f"{_money(resumen.get('ganancia_neta')):,.2f}",
                f"{_pct(resumen.get('roi_porcentaje')):.1f}%",
                f"{_money(resumen.get('ganancia_por_hectarea')):,.2f}",
            ]
            col2 = max(12, max(len(_safe_str(v)) for v in val_samples))
            _set_column_widths(ws_resumen, [min(col1*1.2, 42), min(col2*1.2, 60)])
        except Exception:
            _set_column_widths(ws_resumen, [28, 40])

        # Inversiones
        ws_inv = wb.create_sheet("Inversiones")
        ws_inv.append([
            _woc(ws_inv, "Fecha",      font=header_font, fill=blue_fill),
            _woc(ws_inv, "Categoría",  font=header_font, fill=blue_fill),
            _woc(ws_inv, "Insumos",    font=header_font, fill=blue_fill),
            _woc(ws_inv, "Mano Obra",  font=header_font, fill=blue_fill),
            _woc(ws_inv, "Total",      font=header_font, fill=blue_fill),
            _woc(ws_inv, "Descripción",font=header_font, fill=blue_fill),
        ])
        for inv in reporte_data.get("detalle_inversiones", []) or []:
            ws_inv.append([
                _woc(ws_inv, _first(inv.get("fecha"))),
                _woc(ws_inv, _safe_str(inv.get("categoria") or "Sin categoría")),
                _woc(ws_inv, _money(inv.get("gastos_insumos")),   number_format="#,##0.00"),
                _woc(ws_inv, _money(inv.get("gastos_mano_obra")), number_format="#,##0.00"),
                _woc(ws_inv, _money(inv.get("total")),            number_format="#,##0.00"),
                _woc(ws_inv, _safe_str(inv.get("descripcion"))),
            ])
        ws_inv.freeze_panes = "A2"
        # Autoajuste básico de anchos para inversiones
        try:
            headers = ["Fecha", "Categoría", "Insumos", "Mano Obra", "Total", "Descripción"]
            maxes = [len(h) for h in headers]
            def _len_money(x: Any) -> int:
                try:
                    return len(f"{float(x or 0):,.2f}")
                except Exception:
                    return 4
            for inv in (reporte_data.get("detalle_inversiones") or []):
                maxes[0] = max(maxes[0], len(_first(inv.get("fecha"))))
                maxes[1] = max(maxes[1], len(_safe_str(inv.get("categoria") or "Sin categoría")))
                maxes[2] = max(maxes[2], _len_money(inv.get("gastos_insumos")))
                maxes[3] = max(maxes[3], _len_money(inv.get("gastos_mano_obra")))
                maxes[4] = max(maxes[4], _len_money(inv.get("total")))
                maxes[5] = max(maxes[5], len(_safe_str(inv.get("descripcion"))))
            widths = [min(max(m, 10) * 1.2, 50) for m in maxes]
            _set_column_widths(ws_inv, widths)
        except Exception:
            _set_column_widths(ws_inv, [12, 24, 14, 14, 14, 48])

        # Ventas
        ws_ven = wb.create_sheet("Ventas")
        ws_ven.append([
            _woc(ws_ven, "Fecha",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Variedad",     font=header_font, fill=purple_fill),
            _woc(ws_ven, "Cajas",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Precio/Caja",  font=header_font, fill=purple_fill),
            _woc(ws_ven, "Total",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Gasto",        font=header_font, fill=purple_fill),
            _woc(ws_ven, "Utilidad Neta",font=header_font, fill=purple_fill),
        ])
        neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for v in reporte_data.get("detalle_ventas", []) or []:
            ingreso = _money(v.get("total_venta")); gasto = _money(v.get("gasto"))
            util = _money(v.get("ganancia_neta")) if "ganancia_neta" in v else (ingreso - gasto)
            ws_ven.append([
                _woc(ws_ven, _first(v.get("fecha"))),
                _woc(ws_ven, _safe_str(v.get("tipo_mango"))),
                _woc(ws_ven, int(v.get("num_cajas") or 0),       number_format="0"),
                _woc(ws_ven, _money(v.get("precio_por_caja")),   number_format="#,##0.00"),
                _woc(ws_ven, ingreso,                             number_format="#,##0.00"),
                _woc(ws_ven, gasto,                               number_format="#,##0.00"),
                _woc(ws_ven, util,                                number_format="#,##0.00", fill=(neg_fill if util < 0 else None)),
            ])
        ws_ven.freeze_panes = "A2"
        # Autoajuste básico de anchos para evitar cortes en números grandes
        try:
            det = (reporte_data.get("detalle_ventas") or [])
            def _len_money(x: Any) -> int:
                try:
                    return len(f"{float(x or 0):,.2f}")
                except Exception:
                    return 4
            maxes = [
                len("Fecha"), len("Variedad"), len("Cajas"),
                len("Precio/Caja"), len("Total"), len("Gasto"), len("Utilidad Neta"),
            ]
            for v in det:
                try:
                    maxes[0] = max(maxes[0], len(_first(v.get("fecha"))))
                    maxes[1] = max(maxes[1], len(_safe_str(v.get("tipo_mango"))))
                    maxes[2] = max(maxes[2], len(str(int(v.get("num_cajas") or 0))))
                    maxes[3] = max(maxes[3], _len_money(v.get("precio_por_caja")))
                    maxes[4] = max(maxes[4], _len_money(v.get("total_venta")))
                    maxes[5] = max(maxes[5], _len_money(v.get("gasto")))
                    util = (float(v.get("total_venta") or 0) - float(v.get("gasto") or 0))
                    maxes[6] = max(maxes[6], _len_money(util))
                except Exception:
                    continue
            widths = [min(max(m, 8) * 1.2, 36) for m in maxes]
            _set_column_widths(ws_ven, widths)
        except Exception:
            _set_column_widths(ws_ven, [12, 24, 10, 14, 14, 14, 16])

        # Metadata
        _add_metadata_sheet(wb, reporte_data, "REPORTE DE PRODUCCIÓN - COSECHA INDIVIDUAL")

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generar_excel_temporada(reporte_data: Dict[str, Any]) -> bytes:
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        ws = wb.create_sheet("Resumen Ejecutivo")
        _ws_header(ws, "REPORTE DE PRODUCCIÓN - TEMPORADA COMPLETA", merge_to_col=6)

        info = reporte_data.get("informacion_general", {}) or {}
        res = reporte_data.get("resumen_ejecutivo", {}) or {}

        # Se calcula pero NO se muestra
        fi = _first(info.get("fecha_inicio")); ff = _first(info.get("fecha_fin"))
        periodo_txt = _safe_str(info.get("periodo") or (f"{fi} - {ff}" if (fi or ff) else ""))

        row = 3
        _ws_section_title(ws, row, "INFORMACIÓN GENERAL")
        info_rows = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Ubicación", _safe_str(info.get("ubicacion"))),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Temporada", _safe_str(info.get("temporada_año"))),
            ("Período", periodo_txt),
            ("Estado", _safe_str(info.get("estado"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
            ("Total Cosechas", _safe_str(info.get("total_cosechas"))),
        ]
        # Remueve 'Período' del resumen general de temporada (solo cosecha lo usa)
        info_rows = [it for it in info_rows if not (str(it[0]).strip().lower().startswith('per'))]
        for label, value in info_rows:
            row += 1
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value)])

        row += 2
        _ws_section_title(ws, row, "RESUMEN EJECUTIVO")
        res_rows = [
            ("Inversión Total", _money(res.get("inversion_total")), "#,##0.00"),
            ("Ventas Totales", _money(res.get("ventas_totales")), "#,##0.00"),
            ("Gastos de Venta", _money(res.get("total_gastos_venta")), "#,##0.00"),
            ("Ventas Netas", _money(res.get("ventas_netas")), "#,##0.00"),
            ("Ganancia Neta", _money(res.get("ganancia_neta")), "#,##0.00"),
            ("ROI Temporada (%)", _pct(res.get("roi_temporada")), "0.0"),
            ("Productividad (cajas/ha)", _money(res.get("productividad")), "#,##0.00"),
            ("Cajas Totales", _money(res.get("cajas_totales")), "0"),
        ]
        for label, value, fmt in res_rows:
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value, number_format=fmt)])
        ws.freeze_panes = "A4"
        # Autoajuste básico de anchos para hoja Resumen Ejecutivo (2 columnas)
        try:
            def _len_any(x: Any) -> int:
                try:
                    s = _safe_str(x)
                    return len(s)
                except Exception:
                    return 8
            labels = [
                "Huerta", "Ubicación", "Propietario", "Temporada", "Estado", "Hectáreas", "Total Cosechas",
                "Inversión Total", "Ventas Totales", "Gastos de Venta", "Ventas Netas", "Ganancia Neta",
                "ROI Temporada (%)", "Productividad (cajas/ha)", "Cajas Totales",
            ]
            col1 = max(12, max(_len_any(x) for x in labels))
            # estimado para valores
            col2 = 12
            try:
                vals = [
                    fi, ff, info.get("huerta_nombre"), info.get("ubicacion"), info.get("propietario"),
                    f"{_money(res.get('inversion_total')):,.2f}", f"{_money(res.get('ventas_totales')):,.2f}",
                    f"{_money(res.get('total_gastos_venta')):,.2f}", f"{_money(res.get('ventas_netas')):,.2f}",
                    f"{_money(res.get('ganancia_neta')):,.2f}", f"{_pct(res.get('roi_temporada')):.1f}%",
                    f"{_money(res.get('productividad')):,.2f}", f"{_money(res.get('cajas_totales')):,.0f}",
                ]
                col2 = max(12, max(len(_safe_str(v)) for v in vals))
            except Exception:
                pass
            _set_column_widths(ws, [min(col1*1.1, 40), min(col2*1.1, 40)])
        except Exception:
            _set_column_widths(ws, [28, 24])

        # Comparativo
        ws_comp = wb.create_sheet("Comparativo Cosechas")
        ws_comp.append([
            _woc(ws_comp, "Cosecha",  font=header_font, fill=blue_fill),
            _woc(ws_comp, "Inversión",font=header_font, fill=blue_fill),
            _woc(ws_comp, "Ventas",   font=header_font, fill=blue_fill),
            _woc(ws_comp, "Ganancia", font=header_font, fill=blue_fill),
            _woc(ws_comp, "ROI (%)",  font=header_font, fill=blue_fill),
            _woc(ws_comp, "Cajas",    font=header_font, fill=blue_fill),
        ])
        for c in (reporte_data.get("comparativo_cosechas") or []):
            ws_comp.append([
                _woc(ws_comp, _safe_str(c.get("nombre"))),
                _woc(ws_comp, _money(c.get("inversion")), number_format="#,##0.00"),
                _woc(ws_comp, _money(c.get("ventas")),    number_format="#,##0.00"),
                _woc(ws_comp, _money(c.get("ganancia")),  number_format="#,##0.00"),
                _woc(ws_comp, _pct(c.get("roi")),         number_format="0.0"),
                _woc(ws_comp, _money(c.get("cajas")),     number_format="0"),
            ])
        ws_comp.freeze_panes = "A2"
        # Autoajuste básico comparativo
        try:
            comp = (reporte_data.get("comparativo_cosechas") or [])
            headers = ["Cosecha", "Inversión", "Ventas", "Ganancia", "ROI (%)", "Cajas"]
            maxes = [len(h) for h in headers]
            def _len_money(x: Any) -> int:
                try:
                    return len(f"{float(x or 0):,.2f}")
                except Exception:
                    return 4
            for c in comp:
                maxes[0] = max(maxes[0], len(_safe_str(c.get("nombre"))))
                maxes[1] = max(maxes[1], _len_money(c.get("inversion")))
                maxes[2] = max(maxes[2], _len_money(c.get("ventas")))
                maxes[3] = max(maxes[3], _len_money(c.get("ganancia")))
                try:
                    maxes[4] = max(maxes[4], len(f"{float(c.get('roi') or 0):.1f}%"))
                except Exception:
                    pass
                maxes[5] = max(maxes[5], len(_safe_str(c.get("cajas"))))
            widths = [min(max(m, 10) * 1.2, 36) for m in maxes]
            _set_column_widths(ws_comp, widths)
        except Exception:
            _set_column_widths(ws_comp, [28, 18, 18, 18, 12, 12])

        # Metadata
        _add_metadata_sheet(wb, reporte_data, "REPORTE DE PRODUCCIÓN - TEMPORADA COMPLETA")

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generar_excel_perfil_huerta(reporte_data: Dict[str, Any]) -> bytes:
        wb = Workbook(write_only=True)

        header_font = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        orange_fill = PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid")

        ws = wb.create_sheet("Resumen Histórico")
        _ws_header(ws, "PERFIL HISTÓRICO DE HUERTA", merge_to_col=7)

        info = reporte_data.get("informacion_general", {}) or {}
        row = 3
        _ws_section_title(ws, row, "INFORMACIÓN GENERAL")
        info_rows = [
            ("Huerta", f"{_safe_str(info.get('huerta_nombre'))} ({_safe_str(info.get('huerta_tipo'))})"),
            ("Propietario", _safe_str(info.get("propietario"))),
            ("Ubicación", _safe_str(info.get("ubicacion"))),
            ("Hectáreas", f"{_money(info.get('hectareas')):.2f} ha"),
            ("Variedades", _safe_str(info.get("variedades"))),
            ("Años de Operación", _safe_str(info.get("años_operacion"))),
            ("Temporadas Analizadas", _safe_str(info.get("temporadas_analizadas"))),
        ]
        for label, value in info_rows:
            row += 1
            ws.append([_woc(ws, label, font=Font(bold=True)), _woc(ws, value)])

        row += 2
        _ws_section_title(ws, row, "RESUMEN HISTÓRICO")
        ws.append([
            _woc(ws, "Año",           font=header_font, fill=blue_fill),
            _woc(ws, "Inversión",     font=header_font, fill=blue_fill),
            _woc(ws, "Ventas",        font=header_font, fill=blue_fill),
            _woc(ws, "Ganancia",      font=header_font, fill=blue_fill),
            _woc(ws, "ROI (%)",       font=header_font, fill=blue_fill),
            _woc(ws, "Productividad", font=header_font, fill=blue_fill),
            _woc(ws, "Cosechas",      font=header_font, fill=blue_fill),
        ])
        for d in (reporte_data.get("resumen_historico") or []):
            ws.append([
                _woc(ws, _safe_str(d.get("año"))),
                _woc(ws, _money(d.get("inversion")),     number_format="#,##0.00"),
                _woc(ws, _money(d.get("ventas")),        number_format="#,##0.00"),
                _woc(ws, _money(d.get("ganancia")),      number_format="#,##0.00"),
                _woc(ws, _pct(d.get("roi")),             number_format="0.0"),
                _woc(ws, _money(d.get("productividad")), number_format="#,##0.00"),
                _woc(ws, _safe_str(d.get("cosechas_count"))),
            ])
        ws.freeze_panes = "A5"
        # Autoajuste básico para resumen histórico
        try:
            hist = (reporte_data.get("resumen_historico") or [])
            headers = ["Año", "Inversión", "Ventas", "Ganancia", "ROI (%)", "Productividad", "Cosechas"]
            maxes = [len(h) for h in headers]
            def _len_money(x: Any) -> int:
                try:
                    return len(f"{float(x or 0):,.2f}")
                except Exception:
                    return 4
            for d in hist:
                maxes[0] = max(maxes[0], len(_safe_str(d.get("año"))))
                maxes[1] = max(maxes[1], _len_money(d.get("inversion")))
                maxes[2] = max(maxes[2], _len_money(d.get("ventas")))
                maxes[3] = max(maxes[3], _len_money(d.get("ganancia")))
                try:
                    maxes[4] = max(maxes[4], len(f"{float(d.get('roi') or 0):.1f}%"))
                except Exception:
                    pass
                maxes[5] = max(maxes[5], _len_money(d.get("productividad")))
                maxes[6] = max(maxes[6], len(_safe_str(d.get("cosechas_count"))))
            widths = [min(max(m, 8) * 1.2, 36) for m in maxes]
            _set_column_widths(ws, widths)
        except Exception:
            _set_column_widths(ws, [10, 16, 16, 16, 12, 18, 12])

        # Proyecciones
        pr = reporte_data.get("proyecciones") or {}
        ws2 = wb.create_sheet("Proyecciones")
        _ws_header(ws2, "Proyecciones")
        ws2.append([_woc(ws2, "Métrica", font=header_font, fill=orange_fill), _woc(ws2, "Valor", font=header_font, fill=orange_fill)])
        ws2.append([_woc(ws2, "Proyección Ventas Próx. Temporada"), _woc(ws2, _money(pr.get("proyeccion_proxima_temporada")), number_format="#,##0.00")])
        ws2.append([_woc(ws2, "ROI Esperado (%)"), _woc(ws2, _pct(pr.get("roi_esperado")), number_format="0.0")])
        ws2.append([_woc(ws2, "Recomendaciones"), _woc(ws2, ", ".join([_safe_str(x) for x in (pr.get("recomendaciones") or [])]) or "-")])
        ws2.append([_woc(ws2, "Alertas"), _woc(ws2, ", ".join([_safe_str(x) for x in (pr.get("alertas") or [])]) or "-")])
        ws2.freeze_panes = "A3"
        _set_column_widths(ws2, [36, 48])

        # Metadata
        _add_metadata_sheet(wb, reporte_data, "PERFIL HISTÓRICO DE HUERTA")

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
