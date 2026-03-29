from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from gestion_huerta.models import (
    CategoriaInversion,
    CategoriaPreCosecha,
    Cosecha,
    Huerta,
    InversionesHuerta,
    PreCosecha,
    Propietario,
    Temporada,
    Venta,
)


class OperacionCrudYReportesTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            telefono="7000000004",
            password="pass12345",
            nombre="Admin",
            apellido="Operacion",
            role="admin",
            is_staff=True,
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        propietario = Propietario.objects.create(
            nombre="Propietario",
            apellidos="Operacion",
            telefono="5544444444",
            direccion="Direccion",
        )
        self.huerta = Huerta.objects.create(
            nombre="Huerta Operacion",
            ubicacion="Zona D",
            variedades="Ataulfo",
            historial="",
            hectareas=6,
            propietario=propietario,
        )
        self.today = timezone.localdate()

        self.temporada_operativa = Temporada.objects.create(
            **{"a\u00f1o": self.today.year},
            fecha_inicio=self.today - timedelta(days=30),
            estado_operativo=Temporada.EstadoOperativo.OPERATIVA,
            huerta=self.huerta,
        )
        self.cosecha = Cosecha.objects.create(
            nombre="Cosecha Viva",
            temporada=self.temporada_operativa,
            huerta=self.huerta,
            fecha_inicio=timezone.now() - timedelta(days=1),
        )
        self.categoria_inversion = CategoriaInversion.objects.create(nombre="Insumos")

        self.temporada_planificada = Temporada.objects.create(
            **{"a\u00f1o": self.today.year + 1},
            fecha_inicio=self.today + timedelta(days=20),
            estado_operativo=Temporada.EstadoOperativo.PLANIFICADA,
            huerta=self.huerta,
        )
        self.categoria_precosecha = CategoriaPreCosecha.objects.create(nombre="Preparacion")
        self.precosecha = PreCosecha.objects.create(
            temporada=self.temporada_planificada,
            huerta=self.huerta,
            categoria=self.categoria_precosecha,
            fecha=self.temporada_planificada.fecha_inicio - timedelta(days=5),
            gastos_insumos=Decimal("90.00"),
            gastos_mano_obra=Decimal("30.00"),
            descripcion="Preparacion futura",
        )

    def test_inversion_crud_filtros_archivo_restauracion_y_borrado(self):
        create_response = self.client.post(
            "/huerta/inversiones/",
            {
                "categoria_id": self.categoria_inversion.id,
                "cosecha_id": self.cosecha.id,
                "temporada_id": self.temporada_operativa.id,
                "huerta_id": self.huerta.id,
                "fecha": self.today.isoformat(),
                "gastos_insumos": "120.00",
                "gastos_mano_obra": "30.00",
                "descripcion": "Aplicacion",
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        inversion_id = create_response.json()["data"]["inversion"]["id"]

        list_response = self.client.get(
            f"/huerta/inversiones/?temporada={self.temporada_operativa.id}"
            f"&fecha_desde={self.today.isoformat()}&fecha_hasta={self.today.isoformat()}"
        )
        self.assertEqual(list_response.status_code, 200)
        listed_ids = {item["id"] for item in list_response.json()["data"]["results"]}
        self.assertIn(inversion_id, listed_ids)

        delete_active = self.client.delete(f"/huerta/inversiones/{inversion_id}/")
        self.assertEqual(delete_active.status_code, 400)
        self.assertEqual(
            delete_active.json()["notification"]["key"],
            "inversion_debe_estar_archivada",
        )

        archive_response = self.client.post(
            f"/huerta/inversiones/{inversion_id}/archivar/",
            format="json",
        )
        self.assertEqual(archive_response.status_code, 200)

        restore_response = self.client.post(
            f"/huerta/inversiones/{inversion_id}/restaurar/",
            format="json",
        )
        self.assertEqual(restore_response.status_code, 200)

        second_archive = self.client.post(
            f"/huerta/inversiones/{inversion_id}/archivar/",
            format="json",
        )
        self.assertEqual(second_archive.status_code, 200)

        delete_archived = self.client.delete(f"/huerta/inversiones/{inversion_id}/")
        self.assertEqual(delete_archived.status_code, 200)

    def test_venta_crud_filtros_archivo_restauracion_y_borrado(self):
        create_response = self.client.post(
            "/huerta/ventas/",
            {
                "cosecha_id": self.cosecha.id,
                "temporada_id": self.temporada_operativa.id,
                "huerta_id": self.huerta.id,
                "fecha_venta": self.today.isoformat(),
                "tipo_mango": "Ataulfo",
                "num_cajas": 12,
                "precio_por_caja": 45,
                "gasto": 50,
                "descripcion": "Venta inicial",
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        venta_id = create_response.json()["data"]["venta"]["id"]

        list_response = self.client.get(
            f"/huerta/ventas/?temporada={self.temporada_operativa.id}"
            f"&fecha_desde={self.today.isoformat()}&fecha_hasta={self.today.isoformat()}"
        )
        self.assertEqual(list_response.status_code, 200)
        listed_ids = {item["id"] for item in list_response.json()["data"]["results"]}
        self.assertIn(venta_id, listed_ids)

        delete_active = self.client.delete(f"/huerta/ventas/{venta_id}/")
        self.assertEqual(delete_active.status_code, 400)
        self.assertEqual(delete_active.json()["notification"]["key"], "venta_debe_estar_archivada")

        archive_response = self.client.post(
            f"/huerta/ventas/{venta_id}/archivar/",
            format="json",
        )
        self.assertEqual(archive_response.status_code, 200)

        restore_response = self.client.post(
            f"/huerta/ventas/{venta_id}/restaurar/",
            format="json",
        )
        self.assertEqual(restore_response.status_code, 200)

        second_archive = self.client.post(
            f"/huerta/ventas/{venta_id}/archivar/",
            format="json",
        )
        self.assertEqual(second_archive.status_code, 200)

        delete_archived = self.client.delete(f"/huerta/ventas/{venta_id}/")
        self.assertEqual(delete_archived.status_code, 200)

    def test_categoria_inversion_crud_y_guardas_de_uso(self):
        create_response = self.client.post(
            "/huerta/categorias-inversion/",
            {"nombre": "Fertilizantes"},
            format="json",
        )
        self.assertEqual(create_response.status_code, 201)
        categoria_id = create_response.json()["data"]["categoria"]["id"]

        update_response = self.client.patch(
            f"/huerta/categorias-inversion/{categoria_id}/",
            {"nombre": "Fertilizantes Foliares"},
            format="json",
        )
        self.assertEqual(update_response.status_code, 200)

        delete_active = self.client.delete(f"/huerta/categorias-inversion/{categoria_id}/")
        self.assertEqual(delete_active.status_code, 400)
        self.assertEqual(
            delete_active.json()["notification"]["key"],
            "categoria_debe_estar_archivada",
        )

        categoria = CategoriaInversion.objects.get(pk=categoria_id)
        InversionesHuerta.objects.create(
            categoria=categoria,
            fecha=self.today,
            descripcion="Ligada",
            gastos_insumos=Decimal("10.00"),
            gastos_mano_obra=Decimal("5.00"),
            cosecha=self.cosecha,
            temporada=self.temporada_operativa,
            huerta=self.huerta,
        )

        archive_response = self.client.post(
            f"/huerta/categorias-inversion/{categoria_id}/archivar/",
            format="json",
        )
        self.assertEqual(archive_response.status_code, 200)

        delete_used = self.client.delete(f"/huerta/categorias-inversion/{categoria_id}/")
        self.assertEqual(delete_used.status_code, 400)
        self.assertEqual(
            delete_used.json()["notification"]["key"],
            "categoria_con_inversiones",
        )

        unused_response = self.client.post(
            "/huerta/categorias-inversion/",
            {"nombre": "Transporte"},
            format="json",
        )
        self.assertEqual(unused_response.status_code, 201)
        unused_id = unused_response.json()["data"]["categoria"]["id"]

        archive_unused = self.client.post(
            f"/huerta/categorias-inversion/{unused_id}/archivar/",
            format="json",
        )
        self.assertEqual(archive_unused.status_code, 200)

        delete_unused = self.client.delete(f"/huerta/categorias-inversion/{unused_id}/")
        self.assertEqual(delete_unused.status_code, 200)

    def test_reportes_y_modelo_operativo_aislan_precosecha(self):
        inversion = InversionesHuerta.objects.create(
            categoria=self.categoria_inversion,
            fecha=self.today,
            descripcion="Inversion operativa",
            gastos_insumos=Decimal("150.00"),
            gastos_mano_obra=Decimal("50.00"),
            cosecha=self.cosecha,
            temporada=self.temporada_operativa,
            huerta=self.huerta,
        )
        venta = Venta.objects.create(
            fecha_venta=self.today,
            num_cajas=10,
            precio_por_caja=40,
            tipo_mango="Ataulfo",
            descripcion="Venta operativa",
            gasto=25,
            cosecha=self.cosecha,
            temporada=self.temporada_operativa,
            huerta=self.huerta,
        )

        self.cosecha.refresh_from_db()
        self.assertEqual(self.cosecha.total_gastos, inversion.gastos_totales)
        self.assertEqual(self.cosecha.total_ventas, Decimal(str(venta.total_venta)))
        self.assertEqual(
            self.cosecha.ganancia_neta,
            Decimal(str(venta.total_venta)) - inversion.gastos_totales,
        )

        cosecha_report = self.client.post(
            "/huerta/reportes/cosecha/",
            {
                "cosecha_id": self.cosecha.id,
                "formato": "json",
                "force_refresh": True,
            },
            format="json",
        )
        self.assertEqual(cosecha_report.status_code, 200)
        cosecha_payload = cosecha_report.json()["data"]["reporte"]
        self.assertNotIn("precosechas", cosecha_payload)

        operativa_report = self.client.post(
            "/huerta/reportes/temporada/",
            {
                "temporada_id": self.temporada_operativa.id,
                "formato": "json",
                "force_refresh": True,
            },
            format="json",
        )
        self.assertEqual(operativa_report.status_code, 200)
        operativa_payload = operativa_report.json()["data"]["reporte"]
        self.assertEqual(operativa_payload["precosechas"]["total"], 0.0)

    def test_exportes_de_temporada_y_cosecha_generan_archivos_validos(self):
        temporada_excel = self.client.post(
            "/huerta/reportes/temporada/",
            {
                "temporada_id": self.temporada_planificada.id,
                "formato": "xlsx",
                "force_refresh": True,
            },
            format="json",
        )
        self.assertEqual(temporada_excel.status_code, 200)
        self.assertEqual(
            temporada_excel["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(temporada_excel.content))
        self.assertIn("PreCosecha", workbook.sheetnames)

        cosecha_pdf = self.client.post(
            "/huerta/reportes/cosecha/",
            {
                "cosecha_id": self.cosecha.id,
                "formato": "pdf",
                "force_refresh": True,
            },
            format="json",
        )
        self.assertEqual(cosecha_pdf.status_code, 200)
        self.assertEqual(cosecha_pdf["Content-Type"], "application/pdf")
        self.assertTrue(cosecha_pdf.content.startswith(b"%PDF"))
