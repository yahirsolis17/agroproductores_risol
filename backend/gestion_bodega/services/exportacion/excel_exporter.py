"""Exportacion Excel para reportes de bodega."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


_TITLE_FONT = Font(bold=True, size=16)
_SECTION_FONT = Font(bold=True, size=12)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1A472A", end_color="1A472A", fill_type="solid")


def _safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def _normalize_sheet_name(raw: str, fallback: str) -> str:
    safe = "".join(ch for ch in _safe_str(raw) if ch not in '[]:*?/\\')
    safe = safe.strip() or fallback
    return safe[:31]


def _flatten_dict(prefix: str, value: Any) -> List[Tuple[str, Any]]:
    if isinstance(value, dict):
        out: List[Tuple[str, Any]] = []
        for key, nested in value.items():
            new_prefix = f"{prefix}.{key}" if prefix else _safe_str(key)
            out.extend(_flatten_dict(new_prefix, nested))
        return out
    return [(prefix, value)]


def _append_rows_and_autowidth(ws, rows: Sequence[Sequence[Any]]) -> None:
    col_sizes: Dict[int, int] = {}
    for row in rows:
        ws.append(list(row))
        for idx, cell in enumerate(row, start=1):
            text = _safe_str(cell)
            if len(text) > col_sizes.get(idx, 0):
                col_sizes[idx] = min(max(len(text) + 2, 10), 64)
    for idx, width in col_sizes.items():
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = width


def _write_table_sheet(wb: Workbook, key: str, table: Dict[str, Any], index: int) -> None:
    ws = wb.create_sheet(_normalize_sheet_name(f"Tabla_{index}_{key}", f"Tabla_{index}"))
    ws.append([f"Tabla: {key}"])
    ws["A1"].font = _SECTION_FONT
    ws.append([])

    columns = table.get("columns") or []
    rows = table.get("rows") or []

    if not columns:
        ws.append(["Sin estructura de columnas"])
        return

    ws.append(list(columns))
    for col in range(1, len(columns) + 1):
        c = ws.cell(row=3, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    col_sizes: Dict[int, int] = {idx + 1: min(max(len(_safe_str(col)) + 2, 10), 64) for idx, col in enumerate(columns)}
    if not rows:
        ws.append(["Sin registros"])
        return

    for r in rows:
        values = list(r)
        ws.append(values)
        for idx, value in enumerate(values, start=1):
            size = min(max(len(_safe_str(value)) + 2, 10), 64)
            if size > col_sizes.get(idx, 0):
                col_sizes[idx] = size

    for idx, width in col_sizes.items():
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = width


def _build_excel(reporte_data: Dict[str, Any], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    ws.append([title])
    ws["A1"].font = _TITLE_FONT
    ws.append([])

    metadata = reporte_data.get("metadata") or {}
    rango = reporte_data.get("rango") or {}
    totales = reporte_data.get("totales") or {}
    kpis = reporte_data.get("kpis") or []
    tablas = reporte_data.get("tablas") or {}
    series = reporte_data.get("series") or []

    rows: List[List[Any]] = []
    rows.append(["Metadata"])
    for k, v in _flatten_dict("", metadata):
        rows.append([k, _safe_str(v)])
    if rango:
        rows.append([])
        rows.append(["Rango"])
        for k, v in _flatten_dict("", rango):
            rows.append([k, _safe_str(v)])

    if totales:
        rows.append([])
        rows.append(["Totales"])
        for k, v in _flatten_dict("", totales):
            rows.append([k, _safe_str(v)])

    if kpis:
        rows.append([])
        rows.append(["KPIs"])
        rows.append(["ID", "Label", "Value", "Icon"])
        for kpi in kpis:
            rows.append(
                [
                    _safe_str(kpi.get("id")),
                    _safe_str(kpi.get("label")),
                    _safe_str(kpi.get("value")),
                    _safe_str(kpi.get("icon")),
                ]
            )

    _append_rows_and_autowidth(ws, rows)

    if tablas:
        for idx, (key, table) in enumerate(tablas.items(), start=1):
            if isinstance(table, dict):
                _write_table_sheet(wb, key, table, idx)

    if series:
        ws_series = wb.create_sheet("Series")
        ws_series.append(["Serie ID", "Label", "Type", "X/Name", "Y/Value"])
        for col in range(1, 6):
            c = ws_series.cell(row=1, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center")

        col_sizes = {1: 10, 2: 10, 3: 10, 4: 10, 5: 10}
        for serie in series:
            sid = _safe_str(serie.get("id"))
            slabel = _safe_str(serie.get("label"))
            stype = _safe_str(serie.get("type"))
            for point in serie.get("data", []):
                if not isinstance(point, dict):
                    continue
                x_value = point.get("x", point.get("name", ""))
                y_value = point.get("y", point.get("value", ""))
                row = [sid, slabel, stype, _safe_str(x_value), _safe_str(y_value)]
                ws_series.append(row)
                for idx, value in enumerate(row, start=1):
                    size = min(max(len(_safe_str(value)) + 2, 10), 64)
                    if size > col_sizes.get(idx, 0):
                        col_sizes[idx] = size
        for idx, width in col_sizes.items():
            ws_series.column_dimensions[chr(ord("A") + idx - 1)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


class ExcelExporter:
    @staticmethod
    def generar_excel_semanal(reporte_data: Dict[str, Any]) -> bytes:
        return _build_excel(reporte_data, "Reporte Semanal de Bodega")

    @staticmethod
    def generar_excel_temporada(reporte_data: Dict[str, Any]) -> bytes:
        return _build_excel(reporte_data, "Reporte de Temporada de Bodega")
